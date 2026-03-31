from collections import defaultdict
import csv
from datetime import date, datetime
from io import BytesIO, StringIO

from django.http import HttpResponse
from django.db import IntegrityError, transaction
from rest_framework import permissions, status
from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404

from apps.core.models import Class, Section
from apps.students.models import Student

from .models import StudentAttendance, StudentAttendanceBulk
from .serializers import (
    StudentAttendanceSerializer,
    StudentAttendanceStoreRequestSerializer,
    StudentSearchRequestSerializer,
)
from .services import send_present_attendance_notifications


class AttendanceTenantMixin:
    permission_classes = [permissions.IsAuthenticated]

    def _required_permission_code(self):
        class_name = self.__class__.__name__.lower()
        if "import" in class_name or "downloadsample" in class_name:
            return "student_info.student_attendance_import.view"
        return "student_info.student_attendance.view"

    def initial(self, request, *args, **kwargs):
        super().initial(request, *args, **kwargs)
        user = request.user
        if user.is_superuser:
            return
        code = self._required_permission_code()
        if not hasattr(user, "has_permission_code") or not user.has_permission_code(code):
            raise PermissionDenied("You do not have permission to perform this action.")

    def school_filter(self, request):
        return {} if request.user.is_superuser else {"school_id": request.user.school_id}


class StudentAttendanceListCreateAPIView(AttendanceTenantMixin, APIView):
    def get(self, request):
        queryset = StudentAttendance.objects.select_related("student")
        if not request.user.is_superuser:
            if not request.user.school_id:
                return Response([])
            queryset = queryset.filter(school_id=request.user.school_id)

        params = request.query_params
        if params.get("class_id"):
            queryset = queryset.filter(class_id=params["class_id"])
        if params.get("section_id"):
            queryset = queryset.filter(section_id=params["section_id"])
        if params.get("student_id"):
            queryset = queryset.filter(student_id=params["student_id"])
        if params.get("date"):
            queryset = queryset.filter(attendance_date=params["date"])
        if params.get("month"):
            queryset = queryset.filter(attendance_date__month=params["month"])
        if params.get("year"):
            queryset = queryset.filter(attendance_date__year=params["year"])
        if params.get("academic_year_id"):
            queryset = queryset.filter(academic_year_id=params["academic_year_id"])

        serializer = StudentAttendanceSerializer(queryset.order_by("-attendance_date", "student_id"), many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = StudentAttendanceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            serializer.save(school=request.user.school)
        except IntegrityError:
            return Response(
                {"detail": "Attendance already exists for this student on the selected date."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class StudentAttendanceRetrieveUpdateDeleteAPIView(AttendanceTenantMixin, APIView):
    def get_object(self, request, pk):
        queryset = StudentAttendance.objects.all()
        if not request.user.is_superuser:
            queryset = queryset.filter(school_id=request.user.school_id)
        return get_object_or_404(queryset, pk=pk)

    def get(self, request, pk):
        obj = self.get_object(request, pk)
        return Response(StudentAttendanceSerializer(obj).data)

    def put(self, request, pk):
        obj = self.get_object(request, pk)
        serializer = StudentAttendanceSerializer(obj, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(school=request.user.school)
        return Response(serializer.data)

    def patch(self, request, pk):
        obj = self.get_object(request, pk)
        serializer = StudentAttendanceSerializer(obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save(school=request.user.school)
        return Response(serializer.data)

    def delete(self, request, pk):
        obj = self.get_object(request, pk)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class StudentAttendanceIndexAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP index(): returns classes used in the criteria form."""

    def get(self, request):
        classes = Class.objects.filter(**self.school_filter(request)).order_by("numeric_order", "name")
        return Response([
            {"id": c.id, "class_name": c.name}
            for c in classes
        ])


class StudentSearchAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP studentSearch()."""

    def post(self, request):
        req = StudentSearchRequestSerializer(data=request.data)
        req.is_valid(raise_exception=True)

        class_id = req.validated_data["class"]
        section_id = req.validated_data["section"]
        attendance_date = req.validated_data["attendance_date"]

        students = Student.objects.filter(
            current_class_id=class_id,
            current_section_id=section_id,
            is_active=True,
            **self.school_filter(request),
        ).order_by("id")

        already_assigned_students = []
        new_students = []
        attendance_type = ""

        attendance_rows = {
            row.student_id: row
            for row in StudentAttendance.objects.filter(
                attendance_date=attendance_date,
                class_id=class_id,
                section_id=section_id,
                **self.school_filter(request),
            )
        }

        for student in students:
            attendance = attendance_rows.get(student.id)
            if attendance:
                already_assigned_students.append(
                    {
                        "id": attendance.id,
                        "student_id": attendance.student_id,
                        "attendance_type": attendance.attendance_type,
                        "notes": attendance.notes,
                    }
                )
                attendance_type = attendance.attendance_type
            else:
                new_students.append(
                    {
                        "id": student.id,
                        "admission_no": student.admission_no,
                        "first_name": student.first_name,
                        "last_name": student.last_name,
                        "roll_no": student.roll_no,
                    }
                )

        class_info = Class.objects.filter(id=class_id).first()
        section_info = Section.objects.filter(id=section_id).first()

        # Build student table rows in one payload for frontend parity rendering.
        table_students = []
        for student in students:
            att = attendance_rows.get(student.id)
            table_students.append(
                {
                    "id": student.id,
                    "admission_no": student.admission_no,
                    "first_name": student.first_name,
                    "last_name": student.last_name,
                    "roll_no": student.roll_no,
                    "attendance_type": att.attendance_type if att else None,
                    "attendance_note": att.notes if att else "",
                }
            )

        classes = Class.objects.filter(**self.school_filter(request)).order_by("numeric_order", "name")
        response_data = {
            "classes": [{"id": c.id, "class_name": c.name} for c in classes],
            "date": attendance_date,
            "class_id": class_id,
            "section_id": section_id,
            "attendance_type": attendance_type,
            "already_assigned_students": already_assigned_students,
            "new_students": new_students,
            "students": table_students,
            "search_info": {
                "class_name": class_info.name if class_info else "",
                "section_name": section_info.name if section_info else "",
                "date": attendance_date,
            },
        }
        return Response(response_data)


class StudentAttendanceStoreAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP studentAttendanceStore()."""

    def post(self, request):
        req = StudentAttendanceStoreRequestSerializer(data=request.data)
        req.is_valid(raise_exception=True)
        data = req.validated_data

        attendance_map = data.get("attendance") or data.get("attendance_type") or {}
        note_map = data.get("note") or data.get("attendance_note") or {}

        for student_id in data["id"]:
            existing = StudentAttendance.objects.filter(
                student_id=student_id,
                attendance_date=data["date"],
                **self.school_filter(request),
            ).first()
            if existing:
                existing.delete()

            attendance = StudentAttendance()
            attendance.student_id = student_id
            if data.get("mark_holiday"):
                attendance.attendance_type = "H"
                attendance.notes = "Holiday"
            else:
                raw_type = attendance_map.get(str(student_id))
                if raw_type is None:
                    raw_type = attendance_map.get(student_id)
                attendance.attendance_type = raw_type or "P"

                raw_note = note_map.get(str(student_id))
                if raw_note is None:
                    raw_note = note_map.get(student_id)
                attendance.notes = raw_note or ""

            attendance.attendance_date = data["date"]
            attendance.school = request.user.school
            attendance.academic_year_id = request.data.get("academic_year_id")
            attendance.class_id = request.data.get("class_id")
            attendance.section_id = request.data.get("section_id")
            attendance.save()

            if attendance.attendance_type == "P":
                student = Student.objects.filter(id=student_id, **self.school_filter(request)).select_related("guardian").first()
                if student:
                    send_present_attendance_notifications(student, attendance.attendance_date)

        return Response({"message": "Student attendance been submitted successfully"}, status=status.HTTP_200_OK)


class StudentAttendanceHolidayAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP studentAttendanceHoliday()."""

    def post(self, request):
        class_id = request.data.get("class_id")
        section_id = request.data.get("section_id")
        attendance_date = request.data.get("attendance_date")
        purpose = request.data.get("purpose", "mark")

        students = Student.objects.filter(
            current_class_id=class_id,
            current_section_id=section_id,
            is_active=True,
            **self.school_filter(request),
        )

        if purpose == "mark":
            for student in students:
                existing = StudentAttendance.objects.filter(
                    student_id=student.id,
                    attendance_date=attendance_date,
                    **self.school_filter(request),
                ).first()
                if existing:
                    existing.delete()

                attendance = StudentAttendance(
                    attendance_type="H",
                    notes="Holiday",
                    attendance_date=attendance_date,
                    student_id=student.id,
                    school=request.user.school,
                    academic_year_id=request.data.get("academic_year_id"),
                    class_id=class_id,
                    section_id=section_id,
                )
                attendance.save()

        elif purpose == "unmark":
            for student in students:
                existing = StudentAttendance.objects.filter(
                    student_id=student.id,
                    attendance_date=attendance_date,
                    **self.school_filter(request),
                ).first()
                if existing:
                    existing.delete()

        return Response({"status": "ok"}, status=status.HTTP_200_OK)


class StudentAttendanceMonthlyReportAPIView(AttendanceTenantMixin, APIView):
    def get(self, request):
        class_id = request.query_params.get("class_id")
        section_id = request.query_params.get("section_id")
        month = request.query_params.get("month")
        year = request.query_params.get("year")

        if not class_id or not section_id or not month or not year:
            return Response(
                {"detail": "class_id, section_id, month, year are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        records = (
            StudentAttendance.objects.filter(
                class_id=int(class_id),
                section_id=int(section_id),
                attendance_date__month=int(month),
                attendance_date__year=int(year),
                **self.school_filter(request),
            )
            .select_related("student")
            .order_by("student_id", "attendance_date")
        )

        grouped = defaultdict(lambda: {"P": 0, "A": 0, "L": 0, "F": 0, "H": 0, "name": "", "admission_no": ""})
        for rec in records:
            sid = rec.student_id
            grouped[sid]["name"] = f"{rec.student.first_name} {rec.student.last_name}".strip()
            grouped[sid]["admission_no"] = rec.student.admission_no
            grouped[sid][rec.attendance_type] = grouped[sid].get(rec.attendance_type, 0) + 1

        result = [
            {
                "student_id": sid,
                "admission_no": info["admission_no"],
                "name": info["name"],
                "present": info["P"],
                "absent": info["A"],
                "late": info["L"],
                "half_day": info["F"],
                "holiday": info["H"],
            }
            for sid, info in grouped.items()
        ]
        return Response(result)


class StudentAttendanceImportAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP studentAttendanceImport() for criteria/form data."""

    def get(self, request):
        classes = Class.objects.filter(**self.school_filter(request)).order_by("numeric_order", "name")
        return Response(
            {
                "classes": [{"id": c.id, "class_name": c.name} for c in classes],
            }
        )


class StudentAttendanceDownloadSampleAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP downloadStudentAtendanceFile()."""

    def get(self, request):
        try:
            from openpyxl import Workbook
        except Exception:
            return Response({"detail": "openpyxl is required for sample export."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "student_attendance_sheet"
        headers = ["admission_no", "attendance_date", "attendance_type", "note"]
        sheet.append(headers)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="student_attendance_sheet.xlsx"'
        return response


class StudentAttendanceBulkStoreAPIView(AttendanceTenantMixin, APIView):
    """Parity with PHP studentAttendanceBulkStore()."""

    parser_classes = [MultiPartParser, FormParser]

    def _normalize_date(self, value):
        if value is None or value == "":
            return None
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, (int, float)):
            try:
                from openpyxl.utils.datetime import from_excel

                converted = from_excel(value)
                if isinstance(converted, datetime):
                    return converted.date()
                if isinstance(converted, date):
                    return converted
            except Exception:
                return None
        if isinstance(value, str):
            text = value.strip()
            for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]:
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    def post(self, request):
        attendance_date = request.data.get("attendance_date")
        class_id = request.data.get("class")
        section_id = request.data.get("section")
        uploaded_file = request.FILES.get("file")

        if not attendance_date or not class_id or not section_id or not uploaded_file:
            return Response(
                {"detail": "attendance_date, file, class and section are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ext = uploaded_file.name.split(".")[-1].lower() if "." in uploaded_file.name else ""
        if ext not in {"csv", "xlsx", "xls"}:
            return Response(
                {"detail": "The file must be a file of type: xlsx, csv or xls"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        request_date = self._normalize_date(attendance_date)
        if not request_date:
            return Response({"detail": "Invalid attendance_date."}, status=status.HTTP_400_BAD_REQUEST)

        imported_rows = []

        if ext == "csv":
            content = uploaded_file.read().decode("utf-8", errors="ignore")
            reader = csv.DictReader(StringIO(content))
            for row in reader:
                imported_rows.append(
                    {
                        "admission_no": row.get("admission_no"),
                        "attendance_date": self._normalize_date(row.get("attendance_date")),
                        "attendance_type": (row.get("attendance_type") or "").strip(),
                        "note": row.get("note") or "",
                    }
                )
        else:
            try:
                from openpyxl import load_workbook
            except Exception:
                return Response({"detail": "openpyxl is required for xlsx import."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            workbook = load_workbook(uploaded_file, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return Response({"detail": "File is empty."}, status=status.HTTP_400_BAD_REQUEST)

            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            header_map = {h: i for i, h in enumerate(headers)}

            for row in rows[1:]:
                imported_rows.append(
                    {
                        "admission_no": row[header_map.get("admission_no", -1)] if header_map.get("admission_no", -1) >= 0 else None,
                        "attendance_date": self._normalize_date(row[header_map.get("attendance_date", -1)] if header_map.get("attendance_date", -1) >= 0 else None),
                        "attendance_type": str(row[header_map.get("attendance_type", -1)]).strip() if header_map.get("attendance_type", -1) >= 0 and row[header_map.get("attendance_type", -1)] is not None else "",
                        "note": str(row[header_map.get("note", -1)]).strip() if header_map.get("note", -1) >= 0 and row[header_map.get("note", -1)] is not None else "",
                    }
                )

        school_filter = self.school_filter(request)

        # Legacy importer behavior: stage parsed rows into StudentAttendanceBulk first.
        for value in imported_rows:
            admission_no = (value.get("admission_no") or "").strip()
            if not admission_no:
                continue

            student = Student.objects.filter(admission_no=admission_no, **school_filter).first()
            if not student:
                continue

            row_date = value.get("attendance_date")
            if not row_date:
                continue

            StudentAttendanceBulk.objects.create(
                school=request.user.school,
                student=student,
                attendance_date=row_date,
                attendance_type=(value.get("attendance_type") or "A"),
                note=value.get("note") or "",
                class_id=class_id,
                section_id=section_id,
                student_record_id=None,
            )

        data = list(StudentAttendanceBulk.objects.filter(**school_filter))

        if data:
            class_sections = []
            for value in data:
                if self._normalize_date(value.attendance_date) == request_date:
                    class_sections.append(f"{value.class_id}-{value.section_id}")

            try:
                with transaction.atomic():
                    all_student_ids = []
                    present_students = []

                    for value in sorted(set(class_sections)):
                        class_section = value.split("-")
                        if len(class_section) != 2:
                            continue
                        cs_class_id, cs_section_id = class_section

                        students = Student.objects.filter(
                            current_class_id=cs_class_id,
                            current_section_id=cs_section_id,
                            **school_filter,
                        )

                        for student in students:
                            StudentAttendanceBulk.objects.filter(
                                student_id=student.id,
                                attendance_date=request_date,
                                **school_filter,
                            ).delete()
                            all_student_ids.append(student.id)

                    for value in data:
                        if value is None:
                            continue

                        value_date = self._normalize_date(value.attendance_date)
                        if value_date == request_date:
                            student = Student.objects.filter(id=value.student_id, **school_filter).first()
                            if student:
                                attendance_check = StudentAttendance.objects.filter(
                                    student_id=student.id,
                                    attendance_date=value_date,
                                    **school_filter,
                                ).first()
                                if attendance_check:
                                    attendance_check.delete()

                                present_students.append(student.id)

                                attendance = StudentAttendance()
                                attendance.student_id = student.id
                                attendance.attendance_date = value_date
                                attendance.attendance_type = value.attendance_type
                                attendance.notes = value.note
                                attendance.school = request.user.school
                                attendance.academic_year_id = request.data.get("academic_year_id")
                                attendance.class_id = class_id
                                attendance.section_id = section_id
                                attendance.save()
                        else:
                            # Legacy behavior removes mismatched-date bulk rows by student id.
                            StudentAttendanceBulk.objects.filter(student_id=value.student_id, **school_filter).delete()
            except Exception:
                return Response({"detail": "Operation Failed"}, status=status.HTTP_400_BAD_REQUEST)

        return Response({"message": "Operation successful"}, status=status.HTTP_200_OK)
